# =============================================================
# src/excel_export.py — StationDeck Filled-Template Excel Export
# =============================================================
#
# WHY THIS EXISTS:
#   Stations import their Daily Cash Flow / Stock Movement Excel files,
#   and also enter days directly in the app (OCR daily entry). Those
#   app-entered days exist only in the database — the station's own
#   Excel files fall behind. This module writes the database back INTO
#   the station's real template files, so the exported workbook contains
#   BOTH previously-imported and app-entered data, in the exact layout
#   TotalEnergies expects, and can be re-imported into StationDeck.
#
# HOW (surgical zip editing — NOT load/re-save):
#   Re-saving a workbook with openpyxl rebuilds every part of the file
#   and silently drops what it can't model: embedded charts, cached
#   formula results, zero-width (hidden) columns. Instead, the export
#   copies every zip member of the original byte-for-byte and rewrites
#   ONLY the worksheet XML of sheets that receive new data. Formulas,
#   charts, styles, hidden columns — everything else is untouched by
#   construction.
#
# DATA-SAFETY RULES:
#   1. The master file in data/input/ is never modified — the export is
#      written to data/exports/ under a unique per-request name.
#   2. Rows that already contain data in the master are left byte-perfect.
#      Only rows that are EMPTY in the master but present in the DB
#      (i.e. app-entered days) are filled.
#   3. Filled computed cells keep their formulas — the DB value is written
#      as the cached result, and fullCalcOnLoad makes Excel refresh all
#      totals the first time the file is opened.
#
# WHAT EACH EXPORT CAN FILL:
#   Cash Flow      — every column (the DB stores the full row).
#   Stock Movement — Daily Expenses (11 of 14 categories) and the manual
#                    columns of General Sales Sumary (Lubricants, TBA,
#                    LPG). Fuel dips/purchases and the Shop Sales
#                    day/night split are physical readings the app does
#                    not persist, so those stay as they were imported.
# =============================================================

import logging
import re
import shutil
import zipfile
from datetime import datetime, date
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter

from src.database import get_records_by_date_range

logger = logging.getLogger(__name__)

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# Register every namespace prefix the station templates use, so rewriting a
# worksheet keeps the exact prefixes Excel/WPS expect (ElementTree would
# otherwise invent ns0-style prefixes, which breaks mc:Ignorable references).
for _prefix, _uri in {
    "":     _NS,
    "r":    _R_NS,
    "xdr":  "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "x14":  "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "mc":   "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "etc":  "http://www.wps.cn/officeDocument/2017/etCustomData",
    "xr":   "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
}.items():
    ET.register_namespace(_prefix, _uri)


# ── Cash Flow: DB column → 1-based Excel column on 'CASH FLOW' sheet ─────────
# Mirrors CASHFLOW_COLUMNS in src/reader.py (0-based there, +1 here).
CASHFLOW_INPUT_COLS = {
    "pms_volume":              2,   # B
    "ago_volume":              3,   # C
    "pms_price":               4,   # D
    "ago_price":               5,   # E
    "lubes_litres":            8,   # H
    "lubes_revenue":           9,   # I
    "lpg_kgs":                 10,  # J
    "lpg_revenue":             11,  # K
    "tba_credits":             12,  # L
    "plus_card_payment":       13,  # M
    "shop_sales":              14,  # N
    "plus_card_payment_total": 15,  # O
    "tyre_sales":              16,  # P
    "plus_card_pms":           18,  # R
    "plus_card_ago":           19,  # S
    "other_payments":          20,  # T
    "momo_pay":                21,  # U
    "airtel_pay":              22,  # V
    "visa":                    23,  # W
    "credit_sales":            24,  # X
    "expense_umeme":           26,  # Z
    "expense_water":           27,  # AA
    "expense_security":        28,  # AB
    "expense_stationery":      29,  # AC
    "expense_generator":       30,  # AD
    "expense_meals":           31,  # AE
    "expense_transport":       32,  # AF
    "expense_salaries":        33,  # AG
    "expense_sanitary":        34,  # AH
    "expense_airtime":         35,  # AI
    "expense_misc":            36,  # AJ
    "expense_shop_packaging":  37,  # AK
    "stock_tba":               39,  # AM
    "stock_lpg_acc":           40,  # AN
    "stock_shop_purchase":     41,  # AO
    "actual_cash_banked":      44,  # AR
}

# Computed columns: their formulas stay in place; the DB value is written as
# the cached result so the row is correct immediately and on re-import.
CASHFLOW_COMPUTED_COLS = {
    "pms_revenue":    6,   # F  (=B*D)
    "ago_revenue":    7,   # G  (=C*E)
    "cashless_total": 17,  # Q
    "total_cash":     25,  # Y
    "total_expenses": 38,  # AL
    "cash_to_bank":   43,  # AQ
    "delta":          45,  # AS
}

_CASHFLOW_PRESENCE_COLS = (2, 3, 25)   # B pms_vol, C ago_vol, Y total_cash

# ── Stock Movement: 'Daily Expenses' — DB col → 1-based Excel col ───────────
STOCK_EXPENSE_COLS = {
    "expense_meals":       2,   # B  Meals
    "expense_generator":   3,   # C  Generator
    "expense_umeme":       4,   # D  Electricity
    "expense_water":       5,   # E  Water
    "expense_salaries":    6,   # F  Salaries
    "expense_stationery":  7,   # G  Stationary
    "expense_security":    8,   # H  Security
    "expense_sanitary":    9,   # I  Sanitation
    "expense_airtime":     10,  # J  Airtime/Data
    "expense_transport":   11,  # K  Transport
    "expense_misc":        13,  # M  Sundries
}
STOCK_EXPENSE_TOTAL_COL = 32    # AF — per-row =SUM formula; DB total cached

# 'General Sales Sumary' — only its truly manual columns. PMS/AGO (B/C) and
# Shop (J) are cross-sheet formulas and must not be overwritten.
STOCK_GENERAL_SALES_COLS = {
    "lubes_revenue": 4,   # D  Lubricants
    "tba_credits":   5,   # E  TBA
    "lpg_revenue":   7,   # G  LPG
}


# ─────────────────────────────────────────────────────────────
# Planning helpers (read-only openpyxl)
# ─────────────────────────────────────────────────────────────

def _all_db_records(station_id: str):
    df = get_records_by_date_range("2000-01-01", "2100-01-01", station_id)
    records = {}
    if df is None or df.empty:
        return records
    for _, row in df.iterrows():
        records[str(row["date"])[:10]] = row.to_dict()
    return records


def _date_key(value):
    if isinstance(value, (datetime, date)):
        return f"{value:%Y-%m-%d}"
    return None


def _num(record, key):
    v = record.get(key)
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return int(f) if f == int(f) else round(f, 2)


def _export_path(exports_dir: Path, label: str) -> Path:
    """Unique on-disk name per request. Concurrent exports (double-click,
    browser link-preload) must never share a path — one request overwriting
    the file mid-read corrupts both. Old exports are pruned after 7 days."""
    exports_dir.mkdir(parents=True, exist_ok=True)
    cutoff = datetime.now().timestamp() - 7 * 24 * 3600
    for old in exports_dir.glob("StationDeck_Export_*.xlsx"):
        try:
            if old.stat().st_mtime < cutoff:
                old.unlink()
        except Exception:
            pass
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S_%f")
    return exports_dir / f"StationDeck_Export_{label}_{stamp}.xlsx"


def download_name(label: str) -> str:
    """Clean, date-stamped filename shown to the user."""
    return f"StationDeck_Export_{label}_{datetime.now():%Y-%m-%d}.xlsx"


class _SheetPlan:
    """Read one worksheet (cached values) and decide which cells to fill."""

    def __init__(self, ws_vals):
        self.ws = ws_vals
        self.date_rows = {}           # 'YYYY-MM-DD' -> first row number
        for row in range(2, ws_vals.max_row + 1):
            key = _date_key(ws_vals.cell(row=row, column=1).value)
            if key and key not in self.date_rows:
                self.date_rows[key] = row

    def row_has_any_value(self, row, max_col):
        for col in range(2, max_col + 1):
            v = self.ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)) and v != 0:
                return True
        return False

    def row_is_empty(self, row, presence_cols):
        for col in presence_cols:
            v = self.ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)) and v != 0:
                return False
        return True


# ─────────────────────────────────────────────────────────────
# Surgical zip writer
# ─────────────────────────────────────────────────────────────

def _sheet_part_paths(zf: zipfile.ZipFile) -> dict:
    """Map sheet name -> zip member path (e.g. 'xl/worksheets/sheet6.xml')."""
    wb_xml = zf.read("xl/workbook.xml")
    rels_xml = zf.read("xl/_rels/workbook.xml.rels")
    rid_to_target = {}
    for rel in ET.fromstring(rels_xml):
        rid_to_target[rel.get("Id")] = rel.get("Target")
    result = {}
    root = ET.fromstring(wb_xml)
    for sheet in root.iter(f"{{{_NS}}}sheet"):
        rid = sheet.get(f"{{{_R_NS}}}id")
        target = rid_to_target.get(rid, "")
        if target:
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            result[sheet.get("name")] = target
    return result


def _fmt_num(v) -> str:
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return repr(v)


def _apply_edits_to_sheet_xml(xml_bytes: bytes, edits: dict) -> bytes:
    """Set cell values in a worksheet XML. edits: {(row:int, col:int): number}

    Existing cells keep their style and formula (the new value becomes the
    cached result). Missing cells are created in correct column order,
    borrowing the style of the same column one row above when available.
    """
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{_NS}}}sheetData")
    if sheet_data is None:
        return xml_bytes

    by_row = {}
    for r, c in edits:
        by_row.setdefault(r, {})[c] = edits[(r, c)]

    rows_index = {int(el.get("r")): el for el in sheet_data if el.get("r")}

    def _col_of(cell_el):
        ref = cell_el.get("r") or ""
        letters = "".join(ch for ch in ref if ch.isalpha())
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch) - 64)
        return n

    for row_no, cols in sorted(by_row.items()):
        row_el = rows_index.get(row_no)
        if row_el is None:
            row_el = ET.SubElement(sheet_data, f"{{{_NS}}}row", {"r": str(row_no)})
            rows_index[row_no] = row_el

        cells = {_col_of(c): c for c in row_el.findall(f"{{{_NS}}}c")}

        for col_no, value in sorted(cols.items()):
            ref = f"{get_column_letter(col_no)}{row_no}"
            cell = cells.get(col_no)
            if cell is None:
                cell = ET.Element(f"{{{_NS}}}c", {"r": ref})
                prev = rows_index.get(row_no - 1)
                if prev is not None:
                    for pc in prev.findall(f"{{{_NS}}}c"):
                        if _col_of(pc) == col_no and pc.get("s"):
                            cell.set("s", pc.get("s"))
                            break
                pos = len(list(row_el))
                for i, existing in enumerate(row_el):
                    if _col_of(existing) > col_no:
                        pos = i
                        break
                row_el.insert(pos, cell)
                cells[col_no] = cell

            # Numeric value: drop any string typing / inline string content,
            # keep formulas (Excel recalculates; our value is the cache).
            if cell.get("t"):
                del cell.attrib["t"]
            for child in cell.findall(f"{{{_NS}}}is"):
                cell.remove(child)
            v_el = cell.find(f"{{{_NS}}}v")
            if v_el is None:
                v_el = ET.SubElement(cell, f"{{{_NS}}}v")
            v_el.text = _fmt_num(value)

    return ET.tostring(root, encoding="UTF-8", xml_declaration=True)


def _set_full_calc(workbook_xml: bytes) -> bytes:
    """Ask Excel to recalculate everything on first open, so SUM totals
    pick up the newly filled rows."""
    text = workbook_xml.decode("utf-8")
    if "fullCalcOnLoad" in text:
        return workbook_xml
    new, n = re.subn(r"<calcPr ", '<calcPr fullCalcOnLoad="1" ', text, count=1)
    if n == 0:
        new, n = re.subn(r"(</workbook>)",
                         '<calcPr fullCalcOnLoad="1"/>\\1', text, count=1)
    return new.encode("utf-8")


def _write_export(src: Path, out: Path, sheet_edits: dict) -> None:
    """Copy src → out, rewriting only edited sheets (+ calc flag).

    sheet_edits: {sheet_name: {(row, col): value}}
    """
    with zipfile.ZipFile(src) as zin:
        part_for_sheet = _sheet_part_paths(zin)
        edited_parts = {}
        for sheet_name, edits in sheet_edits.items():
            if not edits:
                continue
            part = part_for_sheet.get(sheet_name)
            if not part:
                logger.warning(f"export: no part found for sheet {sheet_name!r}")
                continue
            edited_parts[part] = _apply_edits_to_sheet_xml(zin.read(part), edits)

        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in edited_parts:
                    data = edited_parts[item.filename]
                elif item.filename == "xl/workbook.xml" and edited_parts:
                    data = _set_full_calc(data)
                zout.writestr(item, data)


# ─────────────────────────────────────────────────────────────
# EXPORT 1 — DAILY CASH FLOW
# ─────────────────────────────────────────────────────────────

def export_cashflow(station_config: dict, station_id: str,
                    exports_dir: Path) -> dict:
    src = Path(station_config["files"]["cashflow"])
    if not src.exists():
        return {"success": False, "path": None, "rows_filled": 0,
                "message": "No cash flow file has been imported yet — "
                           "import one first so StationDeck has the template."}

    records = _all_db_records(station_id)
    if not records:
        return {"success": False, "path": None, "rows_filled": 0,
                "message": "The database has no records to export."}

    wb_vals = openpyxl.load_workbook(src, data_only=True)
    skip = {"Sheet1", "Sheet2", "Sheet3", "Chart1"}
    sheet_edits = {}
    rows_filled = 0

    for sheet_name in wb_vals.sheetnames:
        if sheet_name in skip:
            continue
        plan = _SheetPlan(wb_vals[sheet_name])
        if not plan.date_rows:
            continue
        edits = {}
        for date_str, row_no in plan.date_rows.items():
            record = records.get(date_str)
            if record is None:
                continue
            # Master rows that already hold data are the import source —
            # never overwrite them.
            if not plan.row_is_empty(row_no, _CASHFLOW_PRESENCE_COLS):
                continue
            wrote = False
            for db_col, xl_col in CASHFLOW_INPUT_COLS.items():
                v = _num(record, db_col)
                if v is not None:
                    edits[(row_no, xl_col)] = v
                    wrote = True
            if wrote:
                for db_col, xl_col in CASHFLOW_COMPUTED_COLS.items():
                    v = _num(record, db_col)
                    if v is not None:
                        edits[(row_no, xl_col)] = v
                rows_filled += 1
        if edits:
            sheet_edits[sheet_name] = edits
    wb_vals.close()

    out = _export_path(exports_dir, "Daily_Cash_Flow")
    if sheet_edits:
        _write_export(src, out, sheet_edits)
    else:
        shutil.copy2(src, out)   # nothing to add — perfect copy

    logger.info(f"export_cashflow: {rows_filled} app-entered day(s) "
                f"written into {out.name}")
    return {"success": True, "path": out, "rows_filled": rows_filled,
            "message": f"Cash flow exported — {rows_filled} app-entered "
                       f"day(s) added to the imported data."}


# ─────────────────────────────────────────────────────────────
# EXPORT 2 — STOCK MOVEMENT
# ─────────────────────────────────────────────────────────────

def export_stock(station_config: dict, station_id: str,
                 exports_dir: Path) -> dict:
    src = Path(station_config["files"]["stock"])
    if not src.exists():
        return {"success": False, "path": None, "rows_filled": 0,
                "message": "No stock movement file has been imported yet — "
                           "import one first so StationDeck has the template."}

    records = _all_db_records(station_id)
    if not records:
        return {"success": False, "path": None, "rows_filled": 0,
                "message": "The database has no records to export."}

    wb_vals = openpyxl.load_workbook(src, data_only=True)
    sheet_edits = {}
    rows_filled = 0

    # ── Daily Expenses ────────────────────────────────────────
    if "Daily Expenses" in wb_vals.sheetnames:
        plan = _SheetPlan(wb_vals["Daily Expenses"])
        edits = {}
        for date_str, row_no in plan.date_rows.items():
            record = records.get(date_str)
            if record is None:
                continue
            # Fill ONLY rows that are empty across the ENTIRE row — the
            # sheet has categories beyond what the DB stores (NSSF, VAT,
            # maintenance, cols 15-31), and any of them counts as data.
            if plan.row_has_any_value(row_no, STOCK_EXPENSE_TOTAL_COL):
                continue
            wrote = False
            for db_col, xl_col in STOCK_EXPENSE_COLS.items():
                v = _num(record, db_col)
                if v:   # expenses: only write non-zero, keep sheet sparse
                    edits[(row_no, xl_col)] = v
                    wrote = True
            if wrote:
                v = _num(record, "total_expenses")
                if v:
                    edits[(row_no, STOCK_EXPENSE_TOTAL_COL)] = v
                rows_filled += 1
        if edits:
            sheet_edits["Daily Expenses"] = edits

    # ── General Sales Sumary (manual columns only) ────────────
    if "General Sales Sumary" in wb_vals.sheetnames:
        plan = _SheetPlan(wb_vals["General Sales Sumary"])
        presence = tuple(STOCK_GENERAL_SALES_COLS.values())
        edits = {}
        for date_str, row_no in plan.date_rows.items():
            record = records.get(date_str)
            if record is None:
                continue
            if not plan.row_is_empty(row_no, presence):
                continue
            for db_col, xl_col in STOCK_GENERAL_SALES_COLS.items():
                v = _num(record, db_col)
                if v:
                    edits[(row_no, xl_col)] = v
        if edits:
            sheet_edits["General Sales Sumary"] = edits
    wb_vals.close()

    out = _export_path(exports_dir, "Stock_Movement")
    if sheet_edits:
        _write_export(src, out, sheet_edits)
    else:
        shutil.copy2(src, out)

    logger.info(f"export_stock: {rows_filled} expense day(s) "
                f"written into {out.name}")
    return {"success": True, "path": out, "rows_filled": rows_filled,
            "message": f"Stock movement exported — {rows_filled} app-entered "
                       f"expense day(s) added. Fuel dips and the shop "
                       f"day/night split stay as imported (physical readings "
                       f"the app doesn't capture)."}
