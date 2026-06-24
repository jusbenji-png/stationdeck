"""
reader_manager.py
-----------------
Reads all advanced operational and financial data from the
End-of-Month Manager's Report Excel file.

Provides these functions:
  - read_pnl()                  : Profit & Loss statement (income, expenses, net profit)
  - read_debtors()              : Debtors ledger (customer balances, DSO, overdue flags)
  - read_depositors()           : Pre-payment/depositor balances per customer
  - read_financial_position()   : Full balance sheet (assets, liabilities, net position)
  - read_claims()               : Promotion and fuel claims outstanding
  - read_money_meters()         : Pump meter throughput summary
  - get_all_manager_metrics()   : Single call — returns everything as one dict
"""

import pandas as pd
from pathlib import Path
from datetime import datetime


# ─────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────

def _validate_path(filepath: str | Path) -> Path:
    p = Path(filepath)
    if not p.exists():
        raise FileNotFoundError(f"Manager report not found: {p}")
    return p


def _safe_float(val) -> float:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _load_sheet(wb, sheet_name: str):
    """Load a sheet and return all non-empty rows as a list of tuples."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]
    return [
        row for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]


# ─────────────────────────────────────────────────────────────
# FUNCTION 1 — PROFIT & LOSS
# ─────────────────────────────────────────────────────────────

def read_pnl(filepath: str | Path) -> dict:
    """
    Reads the Profit & Loss statement from the PNL sheet.

    Returns a dict with:
      station_info     : { station, sap_code, profit_center, control_date }
      income_lines     : list of { item, quantity, turnover, margin_ugx, gross_margin }
      gross_income     : float
      expense_lines    : list of { item, amount }
      total_expenses   : float
      net_profit       : float
      price_change_effect : float
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "PNL")

    result = {
        "station_info": {},
        "income_lines": [],
        "gross_income": 0.0,
        "expense_lines": [],
        "total_expenses": 0.0,
        "net_profit": 0.0,
        "price_change_effect": 0.0,
    }

    # --- Extract station metadata (rows 0-10 are header info) ---
    for row in rows:
        label = _safe_str(row[0])
        if label == "STATION":
            result["station_info"]["station"] = _safe_str(row[3])
        elif label == "SAP CODE":
            result["station_info"]["sap_code"] = _safe_str(row[3])
        elif label == "PROFIT CENTER":
            result["station_info"]["profit_center"] = _safe_str(row[3])
        elif label == "DATE ":
            val = row[3]
            if isinstance(val, datetime):
                result["station_info"]["control_date"] = str(val.date())

    # --- Income section ---
    # Income items: col 0 = item name, col 1 = qty, col 2 = turnover,
    #               col 3 = margin per unit, col 4 = gross margin
    income_items = [
        "PMS (ltrs)", "AGO (ltrs)", "Lubricants (ltrs) - New Px",
        "Solar", "LPG Accessories", "LPG(Kgs)", "Shop",
        "Car Wash ", "TBA", "Guichet Unique Commissions ",
    ]

    in_income = False
    in_expenses = False
    total_expenses_captured = False

    for row in rows:
        label = _safe_str(row[0])

        if label == "INCOMES":
            in_income = True
            in_expenses = False
            continue

        if label == "GROSS INCOME":
            result["gross_income"] = _safe_float(row[4])
            in_income = False
            continue

        if label.rstrip() == "EXPENSES":
            in_income = False
            in_expenses = True
            continue

        if label.rstrip() == "TOTAL EXPENSES" and not total_expenses_captured:
            result["total_expenses"] = _safe_float(row[4])
            total_expenses_captured = True
            in_expenses = False
            continue

        if label.rstrip() == "Net Profit/Loss":
            result["net_profit"] = _safe_float(row[4])
            continue

        if label.rstrip() == "RESERVE BALANCE":
            result["reserve_balance"] = _safe_float(row[4])
            continue

        if label.rstrip() == "Price Change effects":
            if result["price_change_effect"] == 0.0:
                result["price_change_effect"] = _safe_float(row[4])
            continue

        if in_income and label in income_items:
            result["income_lines"].append({
                "item":         label.strip(),
                "quantity":     _safe_float(row[1]),
                "turnover_ugx": _safe_float(row[2]),
                "margin_ugx":   _safe_float(row[3]),
                "gross_margin": _safe_float(row[4]),
            })

        if in_expenses and label:
            amount = _safe_float(row[4])
            if amount > 0:
                result["expense_lines"].append({
                    "item":   label.strip(),
                    "amount": amount,
                })

    if "reserve_balance" not in result:
        result["reserve_balance"] = 0.0

    return result


# ─────────────────────────────────────────────────────────────
# FUNCTION 2 — DEBTORS LEDGER
# ─────────────────────────────────────────────────────────────

def read_debtors(filepath: str | Path) -> dict:
    """
    Reads the debtors ledger from the DEBTORS sheet.

    Returns a dict with:
      control_date    : str
      customers       : list of { customer, approved_limit, opening_balance,
                                  consumption, payments, outstanding,
                                  variance_vs_limit, overdue_30d, overdue_60d,
                                  overdue_90d, dso_days }
      total_outstanding : float
      overdue_count   : int  (customers with balance > approved limit)
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "DEBTORS ")

    result = {
        "control_date":    "",
        "customers":       [],
        "total_outstanding": 0.0,
        "overdue_count":   0,
    }

    header_passed = False
    total = 0.0
    overdue = 0

    for row in rows:
        # Control date
        label = _safe_str(row[0])
        if label == "DATE":
            val = row[2]
            if isinstance(val, datetime):
                result["control_date"] = str(val.date())

        # Data rows start after the column header row
        # Header row: DATE | CUSTOMER | APPROVED LIMIT | OPENING BAL | ...
        if _safe_str(row[0]) == "DATE" and _safe_str(row[1]) == "CUSTOMER":
            header_passed = True
            continue

        if not header_passed:
            continue

        customer_name = _safe_str(row[1])
        if not customer_name:
            continue

        outstanding = _safe_float(row[6])
        approved_limit = _safe_float(row[2])

        # Skip zero-balance placeholder rows
        if outstanding == 0.0 and not customer_name:
            continue

        dso_raw = row[11]
        try:
            dso = float(dso_raw) if dso_raw not in (None, "0", 0) else 0.0
        except (ValueError, TypeError):
            dso = 0.0

        customer = {
            "customer":          customer_name,
            "approved_limit":    approved_limit,
            "opening_balance":   _safe_float(row[3]),
            "consumption":       _safe_float(row[4]),
            "payments":          _safe_float(row[5]),
            "outstanding":       outstanding,
            "variance_vs_limit": _safe_float(row[7]),
            "overdue_30d":       _safe_str(row[8]) == "YES",
            "overdue_60d":       row[9] not in (None, 0, 0.0),
            "overdue_90d":       row[10] not in (None, 0, 0.0),
            "dso_days":          round(dso, 1),
        }

        if outstanding > 0:
            total += outstanding
            if outstanding > approved_limit and approved_limit > 0:
                overdue += 1

        result["customers"].append(customer)

    result["total_outstanding"]  = round(total, 2)
    result["overdue_count"]      = overdue
    return result


# ─────────────────────────────────────────────────────────────
# FUNCTION 3 — DEPOSITORS / PRE-PAYMENT CUSTOMERS
# ─────────────────────────────────────────────────────────────

def read_depositors(filepath: str | Path) -> dict:
    """
    Reads pre-payment/depositor balances from the DEPOSITOR-CREDITORS sheet.

    Returns a dict with:
      control_date      : str
      customers         : list of { customer, mode, opening_balance,
                                    period_deposits, consumption, balance }
      total_balance     : float
      active_depositors : int   (customers with balance > 0)
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "DEPOSITOR-CREDITORS-control SH ")

    result = {
        "control_date":      "",
        "customers":         [],
        "total_balance":     0.0,
        "active_depositors": 0,
    }

    header_passed = False
    total = 0.0
    active = 0

    for row in rows:
        label = _safe_str(row[0])
        if label == "DATE":
            val = row[2]
            if isinstance(val, datetime):
                result["control_date"] = str(val.date())

        if _safe_str(row[0]) == "DATE" and _safe_str(row[1]) == "CUSTOMER":
            header_passed = True
            continue

        if not header_passed:
            continue

        customer_name = _safe_str(row[1])
        if not customer_name:
            continue

        balance = _safe_float(row[6])

        customer = {
            "customer":        customer_name,
            "mode":            _safe_str(row[2]),
            "opening_balance": _safe_float(row[3]),
            "period_deposits": _safe_float(row[4]),
            "consumption":     _safe_float(row[5]),
            "balance":         balance,
        }

        total += balance
        if balance > 0:
            active += 1

        result["customers"].append(customer)

    result["total_balance"]     = round(total, 2)
    result["active_depositors"] = active
    return result


# ─────────────────────────────────────────────────────────────
# FUNCTION 4 — FINANCIAL POSITION (BALANCE SHEET)
# ─────────────────────────────────────────────────────────────

def read_financial_position(filepath: str | Path) -> dict:
    """
    Reads the station's financial position (balance sheet view).

    Returns a dict with:
      control_date      : str
      assets            : { stocks: {pms, ago, lubricants, lpg, shop, solar, ...},
                            in_transit: {white_product, ...},
                            other: {trading_balance, uncredited_slips,
                                    cash_at_hand, momo_balance, ...} }
      total_assets      : float
      liabilities       : { uninvoiced_product, unpaid_expenses,
                            depositors_total, ... }
      total_liabilities : float
      net_position      : float
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "FINANCIAL POSITION ")

    result = {
        "control_date":     "",
        "assets": {
            "stocks":     {},
            "in_transit": {},
            "other":      {},
        },
        "total_assets":     0.0,
        "liabilities":      {},
        "total_liabilities":0.0,
        "net_position":     0.0,
    }

    # Asset items mapped from description to key
    stock_items = {
        "PMS":                  "pms",
        "AGO":                  "ago",
        "Lubricants ":          "lubricants",
        "LPG ":                 "lpg",
        "LPG Accessories ":     "lpg_accessories",
        "Shop ":                "shop",
        "Solar ":               "solar",
        "Car Care products ":   "car_care",
        "TBA":                  "tba",
    }
    transit_items = {
        "White Product ":       "white_product",
    }
    other_items = {
        "Trading account balance(if in credit)": "trading_balance",
        "GU Account balance (if in credit)":     "gu_balance",
        " Uncredited slips ":                    "uncredited_slips",
        "Physical Cash at hand":                 "cash_at_hand",
        "Physical Cheques at hand ":             "cheques_at_hand",
        "MOMO pay balance ":                     "momo_balance",
        "Airtel Pay balance ":                   "airtel_balance",
    }
    liability_items = {
        "Uninvoiced Product ":      "uninvoiced_product",
        "Unpaid Expenses ":         "unpaid_expenses",
        "Depositors ":              "depositors",
        "Debtors ":                 "debtors_liability",
        "TOTAL LIABILITIES ":       "_total_liabilities",
        "CURRENT FINANCIAL POSITION": "_net_position",
        "TOTAL ASSETS":             "_total_assets",
    }

    for row in rows:
        label = _safe_str(row[0])
        amount = _safe_float(row[2]) if len(row) > 2 else 0.0

        if label == "DATE":
            val = row[1]
            if isinstance(val, datetime):
                result["control_date"] = str(val.date())

        if label in stock_items:
            result["assets"]["stocks"][stock_items[label]] = amount

        elif label in transit_items:
            result["assets"]["in_transit"][transit_items[label]] = amount

        elif label in other_items:
            result["assets"]["other"][other_items[label]] = amount

        elif label in liability_items:
            key = liability_items[label]
            if key == "_total_liabilities":
                result["total_liabilities"] = amount
            elif key == "_net_position":
                result["net_position"] = amount
            elif key == "_total_assets":
                result["total_assets"] = amount
            else:
                result["liabilities"][key] = amount

    # If totals weren't found as explicit rows, calculate them
    if result["total_assets"] == 0.0:
        all_assets = (
            sum(result["assets"]["stocks"].values()) +
            sum(result["assets"]["in_transit"].values()) +
            sum(result["assets"]["other"].values())
        )
        result["total_assets"] = round(all_assets, 2)

    if result["net_position"] == 0.0 and result["total_assets"] > 0:
        result["net_position"] = round(
            result["total_assets"] - result["total_liabilities"], 2
        )

    return result


# ─────────────────────────────────────────────────────────────
# FUNCTION 5 — PROMOTION & FUEL CLAIMS
# ─────────────────────────────────────────────────────────────

def read_claims(filepath: str | Path) -> dict:
    """
    Reads outstanding promotion and fuel claims from the CLAIMS sheet.

    Returns a dict with:
      control_date   : str
      claims         : list of { description, product, quantity, value_ugx }
      total_claims   : float
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "CLAIMS ")

    result = {
        "control_date": "",
        "claims":       [],
        "total_claims": 0.0,
    }

    header_passed = False
    total = 0.0

    for row in rows:
        label = _safe_str(row[0])

        if label == "DATE":
            val = row[2]
            if isinstance(val, datetime):
                result["control_date"] = str(val.date())

        if _safe_str(row[0]) == "DATE" and _safe_str(row[1]) == "PRODUCT":
            header_passed = True
            continue

        if not header_passed:
            continue

        description = _safe_str(row[0])
        if not description:
            continue

        # Skip summary/total rows
        if description.upper() in ("TOTAL", "GRAND TOTAL", "SUB TOTAL"):
            continue

        value = _safe_float(row[3])
        if value == 0.0:
            continue

        result["claims"].append({
            "description": description,
            "product":     _safe_str(row[1]),
            "quantity":    _safe_str(row[2]),
            "value_ugx":   value,
        })
        total += value

    result["total_claims"] = round(total, 2)
    return result


# ─────────────────────────────────────────────────────────────
# FUNCTION 6 — MONEY METER THROUGHPUT
# ─────────────────────────────────────────────────────────────

def read_money_meters(filepath: str | Path) -> dict:
    """
    Reads pump meter throughput summary from the MONEY METERS sheet.

    Returns a dict with:
      period          : str  (e.g. "05/05/2025 - 07/07/2025")
      products        : list of { product, price, volume_ltrs, amount_ugx }
      total_volume_ltrs : float
      total_amount_ugx  : float
      pumps           : list of { pump_id, opening, closing, sales_ltrs,
                                  price, amount_mm, amount_calc }
    """
    path = _validate_path(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = _load_sheet(wb, "MONEY METERS")

    result = {
        "period":           "",
        "products":         [],
        "total_volume_ltrs": 0.0,
        "total_amount_ugx":  0.0,
        "pumps":            [],
    }

    # Period is in row 0, col 4
    if rows:
        period_val = rows[0][4] if len(rows[0]) > 4 else None
        if period_val:
            result["period"] = _safe_str(period_val).replace("TOTAL SALES: ", "").strip()

    # Product summary block: rows with PDT/PMS/AGO in col 4
    for row in rows:
        col4 = _safe_str(row[4]) if len(row) > 4 else ""
        if col4 in ("PMS", "AGO"):
            product = col4
            price   = _safe_float(row[5]) if len(row) > 5 else 0.0
            volume  = _safe_float(row[6]) if len(row) > 6 else 0.0
            amount  = _safe_float(row[7]) if len(row) > 7 else 0.0
            if volume > 0:
                result["products"].append({
                    "product":    product,
                    "price":      price,
                    "volume_ltrs": volume,
                    "amount_ugx": amount,
                })

        # TOTAL SALES row
        if col4 == "TOTAL SALES":
            result["total_amount_ugx"] = _safe_float(row[7]) if len(row) > 7 else 0.0

    # Pump-level detail: rows with pump IDs like PMS1, PMS2, AGO1, AGO2 etc.
    for row in rows:
        col2 = _safe_str(row[2]) if len(row) > 2 else ""
        if col2 and any(col2.startswith(p) for p in ("PMS", "AGO", "LPG")):
            pump = {
                "pump_id":     col2,
                "opening":     _safe_float(row[3]),
                "closing":     _safe_float(row[4]),
                "sales_ltrs":  _safe_float(row[5]),
                "price":       _safe_float(row[6]),
                "amount_mm":   _safe_float(row[9]),
                "amount_calc": _safe_float(row[10]),
            }
            if pump["sales_ltrs"] > 0:
                result["pumps"].append(pump)

    # Total volume from pump details
    if result["pumps"]:
        result["total_volume_ltrs"] = round(
            sum(p["sales_ltrs"] for p in result["pumps"]), 2
        )

    return result


# ─────────────────────────────────────────────────────────────
# MASTER FUNCTION — ALL METRICS IN ONE CALL
# ─────────────────────────────────────────────────────────────

def get_all_manager_metrics(filepath: str | Path) -> dict:
    """
    Reads all sections from the manager's report in a single call.

    Returns a dict with keys:
      pnl               : output of read_pnl()
      debtors           : output of read_debtors()
      depositors        : output of read_depositors()
      financial_position: output of read_financial_position()
      claims            : output of read_claims()
      money_meters      : output of read_money_meters()

    Each section degrades gracefully — if one fails, the others still load.
    """
    path = _validate_path(filepath)
    result = {}

    sections = {
        "pnl":                read_pnl,
        "debtors":            read_debtors,
        "depositors":         read_depositors,
        "financial_position": read_financial_position,
        "claims":             read_claims,
        "money_meters":       read_money_meters,
    }

    for key, fn in sections.items():
        try:
            result[key] = fn(path)
            print(f"  [OK] {key}")
        except Exception as e:
            print(f"  [WARNING] Could not read '{key}': {e}")
            result[key] = {}

    return result


# ─────────────────────────────────────────────────────────────
# QUICK TEST  (run directly to verify)
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json

    filepath = r"data\input\End of May manager's report.xlsx"

    print("=" * 60)
    print("TEST: get_all_manager_metrics()")
    print("=" * 60)
    metrics = get_all_manager_metrics(filepath)

    for section, data in metrics.items():
        print(f"\n--- {section.upper()} ---")
        print(json.dumps(data, indent=2, default=str))