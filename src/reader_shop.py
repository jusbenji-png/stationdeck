"""
reader_shop.py
--------------
Reads shop daily sales broken down by product category from the
SHOP_MONTHLY_SALES_REPORT Excel file.

Provides two functions:
  - read_shop_daily_sales()    : daily sales per category for a given month/year
  - get_monthly_shop_metrics() : aggregated monthly summary dict for the report

Category columns tracked:
  Non-Alcoholic, Alcoholic, Fresh Products, Ice Cream, Confectionery,
  Tobacco, Grocery, Health & Beauty, Bazaar, Car Care, Car Wash,
  Press, Lottery, Coffee, Snacks, Lubricants in Shop, Guichet, Total
"""

import pandas as pd
from pathlib import Path
from datetime import datetime


# ─────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────

def _validate_path(filepath: str | Path) -> Path:
    p = Path(filepath)
    if not p.exists():
        raise FileNotFoundError(f"Shop report not found: {p}")
    return p


# Month name → sheet name fragment mapping (matches sheet naming pattern)
MONTH_SHEET_MAP = {
    (1,  2023): "Daily sales",
    (7,  2023): "Daily sales JULY 2023",
    (8,  2023): "DAILY SALES AUGUST 2023",
    (9,  2023): "Daily sales for semp",
    (10, 2023): "DAILY SALES OCT 2023",
    (11, 2023): "DAILY SALES NOV 2023",
    (12, 2023): "DAILY SALES DEC 2023 ",
    (1,  2024): "DAILY SALES JAN 2024  (2)",
    (2,  2024): "DAILY SALES FEB 2024  (3)",
    (3,  2024): "DAILY SALES FOR MARCH",
    (4,  2024): "DAILY SALES FOR APRIL 2024 (2)",
    (5,  2024): "DAILY SALES FOR MAY.2024",
    (6,  2024): "DAILY SALES FOR JUNE.2024 (2)",
    (7,  2024): "DAILY SALES FOR JULY.2024 (3)",
    (8,  2024): "DAILY SALES FOR AUGUST.2024 ",
    (9,  2024): "DAILY SALES FOR SEPTEMB.2024 (2",
    (10, 2024): "DAILY SALES FOR OCT.2024",
    (11, 2024): "DAILY SALES FOR NOV.2024 (2)",
    (12, 2024): "DAILY SALES REPORT DEC.2024",
    (1,  2025): "DAIRY SALES REPORT JAN 2025",
    (2,  2025): "DAILY  SALES REPORT FEB.2025",
    (3,  2025): "DAILY  SALES REPORT MAR.2025",
    (4,  2025): "DAILY SALES REPORT FOR APRIL 20",
    (5,  2025): "DAILY SALES REPORT FOR MAY.2025",
    (6,  2025): "DAILY SALES REPORT FOR JUNE2025",
    (7,  2025): "DAIRY REPORT JULY 2025",
    (8,  2025): "DAIRY REPORT AUG.2025",
    (9,  2025): "DAIRY REPORT SEPT 2025",
    (10, 2025): "DAIRY REPORT OCT.2025",
    (11, 2025): "DAILY REPORT NOV.2025",
    (12, 2025): "DAILY REPORT DEC 2025",
    (1,  2026): "DAILY REPORT JAN 2026",
    (2,  2026): "DAILY REPORT FEB 2026",
    (3,  2026): "DAILY REPORT MAR 2026",
    (4,  2026): "DAILY REPORT APRIL 2026",
}

# Column positions in the daily sheet (0-indexed after the DATE column at 0)
CATEGORY_COLS = {
    "non_alcoholic":       1,
    "alcoholic":           2,
    "fresh_products":      3,
    "ice_cream":           4,
    "confectionery":       5,
    "tobacco":             6,
    "grocery":             7,
    "health_beauty":       8,
    "bazaar":              9,
    "car_care":            10,
    "car_wash":            11,
    "press":               12,
    "lottery":             13,
    "coffee":              14,
    "snacks":              15,
    "lubricants_in_shop":  16,
    "guichet":             17,
    # Col 18 = Total T/O (Excluding Airtime, carwash, lubricants in shop)
    # Col 19 = Global Total
    # NOTE: col 18 is often blank in newer sheets; col 19 is the reliable total
    "total_turnover":      19,
    "global_total":        19,
}

# Day label patterns used as the date column (e.g. "1st", "2nd", ...)
DAY_LABELS = {
    "1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5,
    "6th": 6, "7th": 7, "8th": 8, "9th": 9, "10th": 10,
    "11th": 11, "12th": 12, "13th": 13, "14th": 14, "15th": 15,
    "16th": 16, "17th": 17, "18th": 18, "19th": 19, "20th": 20,
    "21st": 21, "22nd": 22, "23rd": 23, "24th": 24, "25th": 25,
    "26th": 26, "27th": 27, "28th": 28, "29th": 29, "30th": 30, "31st": 31,
}


def _find_sheet_name(wb, month: int, year: int) -> str | None:
    """Find the correct daily sales sheet for a given month/year."""
    # Try direct lookup first
    key = (month, year)
    if key in MONTH_SHEET_MAP:
        candidate = MONTH_SHEET_MAP[key]
        if candidate in wb.sheetnames:
            return candidate

    # Fallback: search sheet names for month/year patterns
    import calendar
    month_name = calendar.month_abbr[month].upper()
    month_full = calendar.month_name[month].upper()
    year_str = str(year)
    year_short = str(year)[2:]

    for sheet in wb.sheetnames:
        s = sheet.upper()
        if "DAILY" in s or "DAIRY" in s:
            has_month = (month_name in s or month_full in s)
            has_year = (year_str in s or year_short in s)
            if has_month and has_year:
                return sheet

    return None


# ─────────────────────────────────────────────────────────────
# FUNCTION 1 — DAILY SHOP SALES BY CATEGORY
# ─────────────────────────────────────────────────────────────

def read_shop_daily_sales(filepath: str | Path, month: int, year: int) -> pd.DataFrame:
    """
    Reads daily shop sales broken down by category for the specified month/year.

    Returns a DataFrame with columns:
      day, date, non_alcoholic, alcoholic, fresh_products, ice_cream,
      confectionery, tobacco, grocery, health_beauty, bazaar, car_care,
      car_wash, press, lottery, coffee, snacks, lubricants_in_shop,
      guichet, total_turnover, global_total
    """
    path = _validate_path(filepath)

    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    sheet_name = _find_sheet_name(wb, month, year)
    if sheet_name is None:
        print(f"  [WARNING] No shop daily sales sheet found for {month}/{year}")
        return pd.DataFrame()

    ws = wb[sheet_name]
    records = []

    for row in ws.iter_rows(values_only=True):
        date_val = row[0] if len(row) > 0 else None

        # Only process rows where col 0 is a day label like "1st", "2nd", etc.
        if not isinstance(date_val, str) or date_val.strip() not in DAY_LABELS:
            continue

        day_num = DAY_LABELS[date_val.strip()]
        try:
            date = pd.Timestamp(year=year, month=month, day=day_num)
        except ValueError:
            continue

        record = {"day": day_num, "date": date}
        for col_name, col_idx in CATEGORY_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[col_name] = float(val) if val is not None and pd.notna(val) else 0.0

        records.append(record)

    if not records:
        print(f"  [WARNING] No daily shop data rows found for {month}/{year} in sheet '{sheet_name}'")
        return pd.DataFrame()

    df = pd.DataFrame(records).sort_values("day").reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────
# FUNCTION 2 — MONTHLY SHOP METRICS AGGREGATION
# ─────────────────────────────────────────────────────────────

def get_monthly_shop_metrics(
    filepath: str | Path,
    month: int,
    year: int,
) -> dict:
    """
    Aggregates shop daily sales for the given month/year into a summary
    dictionary ready for use in the report processor.

    Returns a dict with keys:
      category_totals : { non_alcoholic, alcoholic, fresh_products, ice_cream,
                          confectionery, tobacco, grocery, health_beauty,
                          bazaar, car_care, car_wash, press, lottery, coffee,
                          snacks, lubricants_in_shop, guichet }
      total_turnover  : float  (core shop turnover excl. carwash/lubes/airtime)
      global_total    : float  (all-inclusive total)
      trading_days    : int    (number of days with sales data)
      avg_daily_sales : float  (global_total / trading_days)
      best_day        : dict   { day, date, amount }
      worst_day       : dict   { day, date, amount }
      top_3_categories: list   [ { category, total } ]
    """
    path = _validate_path(filepath)
    df = read_shop_daily_sales(path, month, year)

    if df.empty:
        return {}

    category_cols = [
        "non_alcoholic", "alcoholic", "fresh_products", "ice_cream",
        "confectionery", "tobacco", "grocery", "health_beauty",
        "bazaar", "car_care", "car_wash", "press", "lottery", "coffee",
        "snacks", "lubricants_in_shop", "guichet",
    ]

    category_totals = {
        col: round(df[col].sum(), 2)
        for col in category_cols
        if col in df.columns
    }

    total_turnover = round(df["total_turnover"].sum(), 2)
    global_total   = total_turnover  # both point to col 19 now
    # Trading days = days where any category had sales
    trading_days   = len(df[df["non_alcoholic"] + df["alcoholic"] + df["fresh_products"] > 0])
    avg_daily      = round(global_total / trading_days, 2) if trading_days > 0 else 0.0

    # Best and worst trading days (by total_turnover = col 19 = global total)
    df_nonzero = df[df["total_turnover"] > 0]
    if df_nonzero.empty:
        df_nonzero = df  # fallback

    best_row  = df_nonzero.loc[df_nonzero["total_turnover"].idxmax()]
    worst_row = df_nonzero.loc[df_nonzero["total_turnover"].idxmin()]

    best_day  = {"day": int(best_row["day"]),  "date": str(best_row["date"].date()),  "amount": round(best_row["total_turnover"], 2)}
    worst_day = {"day": int(worst_row["day"]), "date": str(worst_row["date"].date()), "amount": round(worst_row["total_turnover"], 2)}

    # Top 3 categories by revenue
    cat_ranking = sorted(
        [{"category": col, "total": category_totals[col]} for col in category_cols if col in category_totals],
        key=lambda x: x["total"],
        reverse=True,
    )[:3]

    return {
        "category_totals":  category_totals,
        "total_turnover":   total_turnover,
        "global_total":     global_total,
        "trading_days":     trading_days,
        "avg_daily_sales":  avg_daily,
        "best_day":         best_day,
        "worst_day":        worst_day,
        "top_3_categories": cat_ranking,
    }


# ─────────────────────────────────────────────────────────────
# QUICK TEST  (run directly to verify)
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json

    filepath = r"data\input\SHOP MONTHLY SALES REPORT FOR APRIL 2026 T.E. Rwizi .xlsx"

    print("=" * 60)
    print("TEST: read_shop_daily_sales() — April 2026")
    print("=" * 60)
    df = read_shop_daily_sales(filepath, month=4, year=2026)
    print(f"  Rows loaded  : {len(df)}")
    if not df.empty:
        print(f"  Date range   : {df['date'].min().date()} → {df['date'].max().date()}")
        print(f"  Total turnover (sum): UGX {df['total_turnover'].sum():,.0f}")
        print(df[["day", "non_alcoholic", "alcoholic", "fresh_products", "total_turnover"]].head(5).to_string(index=False))

    print()
    print("=" * 60)
    print("TEST: get_monthly_shop_metrics() — April 2026")
    print("=" * 60)
    metrics = get_monthly_shop_metrics(filepath, month=4, year=2026)
    print(json.dumps(metrics, indent=2, default=str))